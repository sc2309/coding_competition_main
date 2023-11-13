from flask import Blueprint, render_template, request, flash
import openpyxl

signup2 = Blueprint('signup2', __name__)

def add_values_to_excel(file_path, sheet_name, values):
    try:
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]

        for row in values:
            sheet.append(row)

        workbook.save(file_path)

    except Exception as e:
        print(f"An error occurred: {e}")

@signup2.route('/signup2', methods=['GET', 'POST'])
def index():
    firstname = str(request.form.get('first_name'))
    lastname = str(request.form.get('last_name'))
    number = str(request.form.get('number'))
    email = str(request.form.get('email'))
    password = str(request.form.get('password'))
    is_veg = str(request.form.get('is_veg')).lower()
    is_non_veg = str(request.form.get('is_non_veg')).lower()
    protein_powder_yes = str(request.form.get('protein_powder_yes')).lower()
    excel_file_path = "data.xlsx"
    sheet_name = "DataSheet"
    new_values = [
        [firstname,lastname,number,email,password,is_veg,is_non_veg,protein_powder_yes]
    ]
    add_values_to_excel(excel_file_path, sheet_name, new_values)

    return render_template('signup2.html')